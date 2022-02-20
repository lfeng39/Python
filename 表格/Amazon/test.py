from json import load
import pandas as pd
import time
from openpyxl import Workbook, load_workbook
import numpy as np
import sqlite3


# t1 = time.time()
# wb = load_workbook('D:/data/Amazon/原始数据/11.xlsx',read_only = True)
# t2 = time.time()

# print('完成memoryData11读取，耗时：',round(t2-t1,2))

# CtoGNewList = []
# for i in range(3,1000):
#     cellValue = wb.active.cell(row=i,column = 3).value
#     # print(cellValue)
#     if type(cellValue) != int and type(cellValue) == str:
#         cellValue = 0
#         # print(cellValue,type(cellValue))
#     elif type(cellValue) != int and type(cellValue) != str:
#         cellValue = int(cellValue)
#         # print(memoryData11.iloc[i,2],type(memoryData11.iloc[i,2]))
#     CtoGNewList.append(cellValue)
#     # print('new',cellValue,type(CtoGNewList[i-3]))
# t6 = time.time()

# print(CtoGNewList)
# print('耗时：',round(t6-t2,2))


# abcToCol = {
#     'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4, 'F': 5, 'G': 6, 'H': 7, 'I': 8, 'J': 9, 'K': 10, 'L': 11, 'M': 12,
#     'N': 13,'O': 14, 'P': 15, 'Q': 16, 'R': 17, 'S': 18, 'T': 19, 'U': 20, 'V': 21, 'W': 22, 'X': 23, 'Y': 24, 'Z': 25,
#     'AA': 26, 'AB': 27, 'AC': 28, 'AD': 29, 'AE': 30, 'AF': 31, 'AG': 32, 'AH': 33, 'AI': 34, 'AJ': 35, 'AK': 36, 'AL': 37, 'AM': 38,
#     'AN': 39,'AO': 40, 'AP': 41, 'AQ': 42, 'AR': 43, 'AS': 44, 'AT': 45, 'AU': 46, 'AV': 47, 'AW': 48, 'AX': 49,
# }
# # print(abcToCol.keys())
# keys = [
#     '序号', 'Search Term_1 搜索词', 'Search Term Cnt 搜索词字数',
#     'Occurrence 出现次数', 'Search Frequency Rank Rate',
#     'Search Frequency Rank 1', 'Search Frequency Rank 2',
#     'Search Frequency Rank 3', 'Trend', 'ThreeMonthSFRChange', 'Match',
#     'G_K_O_1', 'G_K_O_2', 'G_K_O_3', 'X.1.Clicked Asin 1',
#     'X.2.Clicked Asin 1', 'X.3.Clicked Asin 1', 'X.1.Product_1',
#     'X.2.Product_1', 'X.3.Product_1', 'X.1.Click Share_1',
#     'X.2.Click Share_1', 'X.3.Click Share_1', 'X.1.Conversion Share_1',
#     'X.2.Conversion Share_1', 'X.3.Conversion Share_1',
#     'X.1.Clicked Asin 2', 'X.2.Clicked Asin 2', 'X.3.Clicked Asin 2',
#     'X.1.Product_2', 'X.2.Product_2', 'X.3.Product_2', 'X.1.Click Share_2',
#     'X.2.Click Share_2', 'X.3.Click Share_2', 'X.1.Conversion Share_2',
#     'X.2.Conversion Share_2', 'X.3.Conversion Share_2',
#     'X.1.Clicked Asin 3', 'X.2.Clicked Asin 3', 'X.3.Clicked Asin 3',
#     'X.1.Product_3', 'X.2.Product_3', 'X.3.Product_3', 'X.1.Click Share_3',
#     'X.2.Click Share_3', 'X.3.Click Share_3', 'X.1.Conversion Share_3',
#     'X.2.Conversion Share_3', 'X.3.Conversion Share_3'
#     ]
# dfDict = {
#     keys[abcToCol['A']]: [1,'abc'],
#     keys[abcToCol['B']]: [2],
#     keys[abcToCol['C']]: [3],
#     keys[abcToCol['D']]: [4],
#     keys[abcToCol['E']]: [5],
#     keys[abcToCol['F']]: [6]
# }
# dfDict[123] = [1,'abc']
# # print(dfDict)

# c = [1, 1, 2, 3, 3, 3, 3, 4, 4]
# c = [i for i in c if i != 3]
# print(c)

# b = ['a','abc']
# c.extend(b)
# print(c)

spc = ['hello world','hello world','helloworld hi','world hello']
num = spc.count('helloworld')
print(num)