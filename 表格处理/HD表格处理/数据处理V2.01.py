#打开数据总表
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import colors, PatternFill

###初始化#############################################
#字典#################################################
_serial_dict = {
                '序号': '序号', 
                '专业类别': '项目类别', 
                '楼盘名称': '项目公司', 
                '工程项目名称': '项目名称', 
                '中标单位': '中标单位', 
                '计价方式': '单价包干', #不取原表
                '招标方式': '招标方式', 
                '投标家数': '投标家数', 
                '规模': '/', #新增
                '定标总价': '订单金额',
                '定标日期': '定标日期', 
                '合同签订日期': '合同签订日期', 
                '经办人': '责任人',
                '移交监察日期': '暂未移交',#新增
                '经济指标': '/',
                '是否含重点项目三层预交楼':'/'
                }

#keys值列表
keys_ = list(_serial_dict.keys())
#values值列表
values_ = list(_serial_dict.values())

######################################打开表格，读取数据############################################
path_xls_win = 'J:/Python/HD表格处理/data/湖北公司采购物资情况台账(4).xlsx'#★★★★★★这里是你要读取的总表表格，路径及文件名可根据自己的需要自定义★★★★★★
r_xls = pd.read_excel(path_xls_win, sheet_name='采购台账', engine='openpyxl').fillna('///')#, dtype={'要求到货时间': str}
print('1)成功读取数据库...')

#表格行数、列数
rows = r_xls.shape[0]#表格行数
cols = r_xls.shape[1]#表格列数

#列表-数据源列名
xls_cols_name = list(r_xls.columns.values)

cho_date = '2021-01'#★★★★★★这里是你要选择的月份，格式不要改，直接改单引号中的数字即可★★★★★★
cho_col = list(r_xls['定标日期'].values)#★★★★★★这里是你要选择的列名，直接修改单引号中间的文字即可★★★★★★


#新表列名在总表中的位置###############################
cho_cols_index = []
# cho_cols = [0, 19, 4, 10, 11, 12, 13, 14, 20, 7]
for i in range(1, len(values_)):
    if values_[i] in xls_cols_name:
        # print(xls_cols_name.index(values_[i]))
        cho_cols_index.append(xls_cols_name.index(values_[i]))
    else:
        pass
# print(cho_cols_index)
######################################按指定条件，获取到读取的数据行############################################
cho_rows = []#储存需要数据的所在行
#cho_rows = [1938, 1939, 1941, 1942, 1943, 1947, 1948, 1950, 1952, 1953, 1954, 1955, 1956, 1957, 1959, 1962, 1963, 1964, 1965, 1966, 1967, 1968, 1973, 1987, 1988, 1989, 1990, 1991, 3678, 3680, 3697]
for i in range(rows):
    if cho_date in str(cho_col[i]):
        cho_rows.append(i)#将需要数据所在行添加进list
# print(cho_rows)
print('2)成功抓取指定数据行...')

# type_list = ['营销', '工程', '资产']
type_list = list(set(r_xls['项目类别'].values))

for i in range(len(type_list)):
    if type_list[i] == '///':
        type_list.pop(i)
        break

print(type_list)
# print(len(type_list))

#构建类别分组列表
gro_01, gro_02, gro_03, gro_04, gro_05, gro_06, gro_07, gro_08 = [], [], [], [], [], [], [], []

for k in range(len(cho_rows)):

    new_row = []#{}#创建用于储存指定行的指定列的值的列表
    for v in cho_cols_index:
        new_col = r_xls.iloc[cho_rows[k], v]#按照index序列，和选择的行，读取单元格数据

        if v == 14:#将时间列缩减
            new_col = str(new_col)[:10]

        new_row.append(new_col)
    #补齐缺损字段
    new_row.insert(0, str(k+1))
    new_row.insert(5, values_[5])
    new_row.insert(8, values_[8])
    new_row.insert(13, values_[13])
    new_row.insert(14, values_[14])
    new_row.insert(15, values_[15])

    # print(new_row)
    if '工程' in new_row:
        gro_01.append(new_row)
    elif '营销' in new_row:
        gro_02.append(new_row)
    elif '资产' in new_row:
        gro_03.append(new_row)
    elif '运营' in new_row:
        gro_04.append(new_row)
    elif '行政' in new_row:
        gro_05.append(new_row)
    elif '物业' in new_row:
        gro_06.append(new_row)
    elif '高科' in new_row:
        gro_07.append(new_row)
    elif '呈批' in new_row:
        gro_08.append(new_row)

    new_df = gro_01 + gro_02 + gro_03 + gro_04 + gro_05 + gro_06 + gro_07 + gro_08
    # new_df.insert(0,values_)

print('yes')

create_xls_home = pd.DataFrame(new_df, columns= keys_)
# create_xls_home = create_xls_home.set_index('序号')
create_xls_home.to_excel('C:/Users/lfeng/Desktop/cc.xlsx')#★★★★★★这里是创建的表格路径及名称，可自己修改单引号中的内容★★★★★★
# create_xls_home.to_excel('F:/Python/HD表格处理/data/test.xlsx')#★★★★★★这里是创建的表格路径及名称，可自己修改单引号中的内容★★★★★★


# wb = load_workbook('C:/Users/lfeng/Desktop/cc.xlsx')
# ws = wb.active
# ws.cell(row=1, column=1, value=10)
# workbook = openpyxl.Workbook()
# # sheet = workbook.create_sheet(index = 0, title = 'test')
# # sheet.insert_rows(0,'test')

# # for i in range(len(new_df)):
# #     sheet.append(new_df[i])
# # sheet.cell(row=1, column=1).fill = PatternFill("solid", fgColor='ff0000')
# workbook.save('C:/Users/lfeng/Desktop/cc.xlsx')
# print(new_df)
# for v in range(len(new_df)):
#     print(new_df[v])